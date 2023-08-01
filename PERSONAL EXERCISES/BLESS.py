from tabulate import tabulate
from openpyxl import Workbook,load_workbook
import datetime
class Student:
  def __init__(self):
      self.Student_data=[]
      self.get_student_data()
      
    
  def get_student_data(self):
    workbook = load_workbook(filename="bless1.xlsx")
    worksheet = workbook.active
    for row in worksheet.iter_rows(min_row=None,values_only=True):
      name=row[0]
      parent=row[1]
      place_of_birth=row[2]
      dob=row[3]
      phone=row[4]
      mail=row[5]
      student={'name':name,'parent':parent,'place_of_birth':place_of_birth,'dob':dob,'phone':phone,'mail':mail}
      self.Student_data.append(student)
    print("student data already fetched")
    
  def send_student_data(self):
    workbook=Workbook()
    for student_data in self.Student_data:
      sheet=workbook.create_sheet(title=student_data['name'])
      sheet.append(['name','parent','place_of_birth','dob','phone','mail'])
      #dob=student_data['dob']
      #age=self.calculate_age(student_data)
      sheet.append(['AGE',age])
    workbook.save('bless2.xlsx')
    print("student data stored")
    
  # def calculate_age(self,Student_data):
   # dob=Student_data['dob']
   # age = (datetime.datetime.now() - dob).days // 365
   # return age
    
  def display_student_information(self):
    all=("Thank you")
    return all
student=Student()
student.send_student_data()
student.display_student_information()
print("program successful")